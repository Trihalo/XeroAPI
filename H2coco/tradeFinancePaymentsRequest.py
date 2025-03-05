import sys
import os
import json
import pandas as pd
import requests
import logging
from datetime import datetime
from openpyxl import Workbook  # Add this import

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from helpers.fetchInvoicesForClient import fetchInvoicesForClient

if not os.path.exists('logs'): os.makedirs('logs')
log_filename = datetime.now().strftime('logs/payment_%Y%m%d_%H%M%S.log')

# Configure logging without timestamp
logging.basicConfig(filename=log_filename, level=logging.INFO, 
                    format='%(levelname)s - %(message)s')

def create_payment_status_excel(poInvoiceDict, paidPOs, unpaidPOs):
    wb = Workbook()
    ws_paid = wb.create_sheet("Paid POs")
    ws_unpaid = wb.create_sheet("Unpaid POs")

    ws_paid.append(["PO Number", "Supplier Invoice Number"])
    ws_unpaid.append(["PO Number"])

    for poNumber in paidPOs:
        invoiceData = poInvoiceDict[poNumber]
        ws_paid.append([poNumber, invoiceData["supplierInvNumber"]])

    for poNumber in unpaidPOs: ws_unpaid.append([poNumber])

    folder_name = "PO_Payment_Status"
    os.makedirs(folder_name, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{folder_name}/PO_Payment_Status_{timestamp}.xlsx"

    wb.save(file_name)
    logging.info(f"Payment status Excel file has been saved to {file_name}")


def main():
    client = "H2coco"
    invoiceStatus = "AUTHORISED"

    allInvoices = []
    clientTokens = {}

    try:
        invoices, accessToken, xeroTenantId = fetchInvoicesForClient(client, invoiceStatus)
        
        if not isinstance(invoices, list):
            raise Exception(f"Expected a list of invoices but got {type(invoices)} for {client}")
        
        allInvoices.extend(invoices)
        
        clientTokens[client] = {
            "accessToken": accessToken,
            "xeroTenantId": xeroTenantId
        }
    except Exception as e:
        logging.error(f"Error fetching invoices: {e}")
        return

    accpayInvoices = [
        invoice for invoice in allInvoices
        if isinstance(invoice, dict) and invoice.get("Type") == "ACCPAY"
    ]

    # Ask for the file path
    filePath = 'test.xlsx'

    # Read the Excel file to get the PO numbers, dates, currency rates, and amounts
    df = pd.read_excel(filePath)
    poNumbers = df['PO'].tolist()
    dates = df['Date'].dt.strftime('%Y-%m-%d').tolist()  # Convert dates to string format
    currencyRates = df['CurrencyRate'].tolist()
    amounts = df['Amount'].tolist()

    poInvoiceDict = {}
    for poNumber in poNumbers:
        poReference = f"PO-{poNumber}"
        invoiceId = None
        supplierInvNumber = None
        for invoice in accpayInvoices:
            if poReference in invoice.get("InvoiceNumber", ""):
                invoiceId = invoice.get("InvoiceID")
                supplierInvNumber = invoice.get("InvoiceNumber", "").split(' ', 1)[1] if ' ' in invoice.get("InvoiceNumber", "") else ""
                break
        poInvoiceDict[poNumber] = {
            "InvoiceID": invoiceId,
            "supplierInvNumber": supplierInvNumber,
            "InvoiceDate": invoice.get("DateString", "").split('T')[0] if invoice.get("DateString", "") else "",
            "AmountPaid": invoice.get("AmountPaid", 0)
        }

    paidPOs = []
    unpaidPOs = []

    # Create and send individual payment requests
    for poNumber, date, currencyRate, amount in zip(poNumbers, dates, currencyRates, amounts):

        invoiceData = poInvoiceDict.get(poNumber)

        # compare the dates to see which one should be used
        invoiceDate = datetime.strptime(invoiceData["InvoiceDate"], "%Y-%m-%d")
        dateObj = datetime.strptime(date, "%Y-%m-%d")
        paymentDate = invoiceDate if dateObj < invoiceDate else dateObj
        
        if invoiceData and invoiceData["supplierInvNumber"] and invoiceData["AmountPaid"] == 0.0:
            payment = {
                "Invoice": {
                    "InvoiceID": invoiceData["InvoiceID"]
                },
                "Account": {
                    "Code": "2010"
                },
                "Date": paymentDate.strftime("%Y-%m-%d"),
                "CurrencyRate": currencyRate,
                "Amount": amount,
                "Reference": f"PO{poNumber} {invoiceData['supplierInvNumber']} USD {amount}"
            }

            paymentPayload = {
                "Payments": [payment]
            }

            # Send the payment request to the Xero API
            url = "https://api.xero.com/api.xro/2.0/Payments"
            headers = {
                "Authorization": f"Bearer {accessToken}",
                "Content-Type": "application/json",
                "Xero-tenant-id": xeroTenantId
            }

            response = requests.post(url, headers=headers, data=json.dumps(paymentPayload))
            if response.status_code in [200, 201]:
                logging.info(f"Payment for PO {poNumber} ({invoiceData['supplierInvNumber']}) of ${amount} at exchange rate of {currencyRate} allocated on {paymentDate.strftime("%Y-%m-%d")}.")
                paidPOs.append(poNumber)
            else:
                logging.error(f"Failed to allocate payment for PO {poNumber} ({invoiceData['supplierInvNumber']}): {response.status_code} - {response.text}")
                unpaidPOs.append(poNumber)

            logging.info("")
        elif invoiceData and not invoiceData["supplierInvNumber"]:
            logging.error(f"PO {poNumber}'s payment not allocated since supplier invoice number is missing")
            unpaidPOs.append(poNumber)
        elif invoiceData["AmountPaid"] > 0:
            logging.error(f"PO {poNumber} has already has a prepayment allocated to it. Please manually check")
        elif not invoiceData:
            logging.error(f"PO {poNumber} not found in bills")
            unpaidPOs.append(poNumber)

    # Output the dictionary to a JSON file
    with open("poInvoiceMapping.json", "w") as outfile:
        json.dump(poInvoiceDict, outfile, indent=4)

    logging.info("PO to Invoice ID mapping has been saved to poInvoiceMapping.json")

    # Create the Excel file with payment status
    create_payment_status_excel(poInvoiceDict, paidPOs, unpaidPOs)

if __name__ == "__main__":
    main()