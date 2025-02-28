import os
import sys
import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment
import requests
import time

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from helpers.dateStringsHelper import getSydneyDate

# Helper functions

def getCategory(invoice):
    if "LineItems" in invoice and isinstance(invoice["LineItems"], list):
        for line_item in invoice["LineItems"]:
            if "Tracking" in line_item and isinstance(line_item["Tracking"], list):
                for tracking_item in line_item["Tracking"]:
                    if tracking_item.get("Name") == "Category": 
                        return tracking_item.get("Option")
    return None


def getConsultant(invoice):
    consultants = []
    line_items = invoice.get("LineItems", [])

    for line in line_items:
        if "Tracking" in line and isinstance(line["Tracking"], list):
            for tracking_item in line["Tracking"]:
                if tracking_item.get("Name") == "Consultant":
                    consultant_name = tracking_item.get("Option")
                    if consultant_name not in consultants: 
                        consultants.append(consultant_name)

    return ", ".join(consultants) if consultants else "Unknown"


def fetchInvoiceHistory(invoice_id, access_token, xero_tenant_id):
    XERO_HISTORY_API_URL = "https://api.xero.com/api.xro/2.0/Invoices/{invoice_id}/History"

    if not access_token or not xero_tenant_id:
        raise Exception("Missing Xero authentication credentials.")

    url = XERO_HISTORY_API_URL.format(invoice_id=invoice_id)
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json",
        "Xero-Tenant-Id": xero_tenant_id
    }

    response = requests.get(url, headers=headers)

    if response.status_code == 200: 
        return response.json()
    
    elif response.status_code == 429:
        retry_after = int(response.headers.get("Retry-After", 5))  
        print(f"⚠️ Rate limit hit! Retrying in {retry_after} seconds for invoice {invoice_id}...")
        time.sleep(retry_after)
        return fetchInvoiceHistory(invoice_id, access_token, xero_tenant_id) 

    else:
        print(f"❌ Error fetching invoice history: {invoice_id}: {response.status_code} - {response.text}")
        return None


def getNotes(invoice_id, invoice_number, client_tokens):
    """Fetch notes using the correct Xero credentials based on invoice number prefix."""
    
    # Determine which credentials to use
    if invoice_number.startswith("TC"):
        client_key = "FUTUREYOU_CONTRACTING"
    else:
        client_key = "FUTUREYOU_RECRUITMENT"

    if client_key not in client_tokens:
        raise Exception(f"❌ No credentials found for {client_key}")

    access_token = client_tokens[client_key]["access_token"]
    xero_tenant_id = client_tokens[client_key]["xero_tenant_id"]

    notes = []
    history = fetchInvoiceHistory(invoice_id, access_token, xero_tenant_id)

    if history:
        history_records = history.get("HistoryRecords", [])
        for record in history_records:
            if record.get("Changes") == "Note":
                details = record.get("Details", "").strip()
                date_string = record.get("DateUTCString", "")

                # Convert date string to DD/M format
                try:
                    note_date = getSydneyDate(date_string)
                    formatted_date = note_date.strftime("%d/%m")
                except ValueError:
                    formatted_date = "Unknown Date"

                note_entry = f"{formatted_date}: {details}"

                if note_entry not in notes:  
                    notes.append(note_entry)

    print(f"✅ Notes fetched for invoice {invoice_number}")
    time.sleep(1)  # Enforce rate limit (1 request per second)
    return ", ".join(notes) if notes else ""


def processAtbData(data, client_tokens):
    """Process invoices and export to Excel."""
    
    accrec_invoices = [
        invoice for invoice in data.get("Invoices", [])
        if isinstance(invoice, dict) and 
           invoice.get("Status") == "AUTHORISED" and  
           invoice.get("Type") == "ACCREC"
    ]
    
    invoices = []

    for index, invoice in enumerate(accrec_invoices, start=1):
        invoice_id = invoice.get("InvoiceID", "")
        invoice_number = invoice.get("InvoiceNumber", "")

        invoice_date = datetime.strptime(invoice["DateString"], "%Y-%m-%dT%H:%M:%S")
        due_date = datetime.strptime(invoice["DueDateString"], "%Y-%m-%dT%H:%M:%S")

        formatted_invoice_date = invoice_date.strftime("%d/%m/%Y")
        formatted_due_date = due_date.strftime("%d/%m/%Y")

        type_column = ""
        if "Retainer Commencement" in invoice.get("Reference", ""): 
            type_column = "Commencement Retainer"
        if (date.today() - invoice_date.date()).days > 90: 
            type_column = "Invoices 90 days plus"

        # Invoice total
        amount_due = float(invoice.get("AmountDue", 0.0))
        currency_rate = float(invoice.get("CurrencyRate", 1.0))

        # Convert to AUD if needed
        if invoice.get("CurrencyCode", "AUD") != "AUD" and currency_rate != 0:
            amount_due *= (1 / currency_rate)

        print(f"📊 Processing invoice {index}/{len(accrec_invoices)} - {invoice_number}")

        invoices.append({
            "Invoice Number": invoice_number,
            "Type": type_column,
            "Notes": "",
            "Contact": invoice.get("Contact", {}).get("Name", ""),
            "Invoice Date": formatted_invoice_date,
            "Ageing Days": "",
            "Due Date": formatted_due_date,
            "Overdue Days": "",
            "Invoice Number.": invoice_number,
            "Invoice Reference": invoice.get("Reference", ""),
            "Total": amount_due,
            "Category": getCategory(invoice),
            "Consultant": getConsultant(invoice),
            "Comments": getNotes(invoice_id, invoice_number, client_tokens),
        })

    print("✅ All invoices processed. Generating Excel file...")

    df = pd.DataFrame(invoices)
    df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], format="%d/%m/%Y")
    df = df.sort_values(by="Invoice Date", ascending=True)
    df["Invoice Date"] = df["Invoice Date"].dt.strftime("%d/%m/%Y")

    # file name
    date_string = datetime.today().strftime("%d-%m-%Y")
    os.makedirs("ATB", exist_ok=True)
    output_file = os.path.join("ATB", f"ATB {date_string}.xlsx")
    
    df.to_excel(output_file, index=False, startrow=1)

    wb = load_workbook(output_file)
    ws = wb.active

    ws["A1"] = f"{datetime.today().strftime('%d/%m/%Y')}"

    last_row = len(df) + 2

    for row in range(3, last_row + 1):
        ws[f"F{row}"] = f"=$A$1-E{row}"
        ws[f"H{row}"] = f"=$A$1-G{row}"

    ws["K1"] = f"=SUBTOTAL(9, K3:K{last_row})"
    ws["L1"] = f"=SUMIF(A3:A{last_row}, \"?????\", K3:K{last_row})"
    ws["M1"] = f"=SUMIF(A3:A{last_row}, \"TC*\", K3:K{last_row})"

    # formatting 

    accounting_format = NamedStyle(name="accounting_format")
    accounting_format.number_format = '"$"#,##0.00'

    number_format = NamedStyle(name="number_format")
    number_format.number_format = '0'

    date_format = NamedStyle(name="date_format")
    date_format.number_format = "DD/MM/YYYY"

    wrap_alignment = Alignment(wrap_text=True)
    wb.add_named_style(accounting_format)
    wb.add_named_style(number_format)
    wb.add_named_style(date_format)
    
    for row in range(1, last_row + 1):
        if row != 2: 
            ws[f"K{row}"].style = "accounting_format"
            ws[f"E{row}"].style = "date_format"
            ws[f"F{row}"].style = "number_format"
            ws[f"G{row}"].style = "date_format"
            ws[f"H{row}"].style = "number_format"
            ws[f"M{row}"].alignment = wrap_alignment 
            ws[f"N{row}"].alignment = wrap_alignment 

    ws["L1"].style = "accounting_format"
    ws["M1"].style = "accounting_format"


    column_widths = {"A": 15, "B": 25, "C": 20, "D": 40, "E": 12, "F": 12, "G": 12, "H": 12,
                     "I": 15, "J": 40, "K": 15, "L": 12, "M": 25, "N": 70}

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_file)
    print(f"📂 Excel file '{output_file}' created successfully.")

    return output_file
