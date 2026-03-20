#!/usr/bin/env python3
"""
generateTalentMapExcel.py

Reads rawData.csv (a Bullhorn Longlist export), enriches each candidate via
the Bullhorn REST API (location, salary, LinkedIn URL), and produces one
formatted Excel talent map per unique job order.

Usage:
    python generateTalentMapExcel.py                   # uses rawData.csv, outputs to ./output/
    python generateTalentMapExcel.py --csv path.csv    # custom CSV path
    python generateTalentMapExcel.py --no-api          # skip API enrichment (CSV data only)
"""

import csv
import os
import sys
import time
import argparse
from datetime import datetime

from urllib.parse import urlsplit

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Reach repo root to import xeroAuthHelper (get_github_variable / update_github_variable)
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../../..")))
from xeroAuthHelper import get_github_variable, update_github_variable

# ── Config ────────────────────────────────────────────────────────────────────
CLIENT_ID     = os.environ.get("FUTUREYOU_BULLHORN_CLIENT_ID")
CLIENT_SECRET = os.environ.get("FUTUREYOU_BULLHORN_CLIENT_SECRET")
USERNAME      = "futureyou.restapi"
PASSWORD      = os.environ.get("FUTUREYOU_BULLHORN_PASSWORD")
REDIRECT_URI  = "https://welcome.bullhornstaffing.com"

# Once you identify which customText field holds LinkedIn URLs in FutureYou's
# Bullhorn instance, set it here to skip auto-discovery on every run.
# e.g., LINKEDIN_FIELD = "customText3"
LINKEDIN_FIELD = None

# Explicit field list required by Bullhorn API (fields=* is not supported).
# Requesting all scalar/address fields plus the full customText1–40 range so we
# can inspect the raw dump and identify which field holds the LinkedIn URL.
_custom_text   = ",".join(f"customText{i}"      for i in range(1, 41))
_custom_int    = ",".join(f"customInt{i}"        for i in range(1, 24))
_custom_float  = ",".join(f"customFloat{i}"      for i in range(1, 24))
_custom_date   = ",".join(f"customDate{i}"       for i in range(1, 14))
_custom_block  = ",".join(f"customTextBlock{i}"  for i in range(1, 11))

CANDIDATE_FIELDS = ",".join(filter(None, [
    # Identity
    "id,firstName,lastName,middleName,name,namePrefix,nameSuffix,nickName",
    # Contact
    "email,email2,email3,phone,phone2,phone3,mobile,workPhone",
    # Address (returns nested object with city/state/zip etc.)
    "address,secondaryAddress",
    # Employment
    "companyName,companyURL,occupation,employeeType,experience,employmentPreference,skillSet",
    # Compensation
    "salary,salaryLow,hourlyRate,hourlyRateLow,dayRate,dayRateLow",
    # Profile
    "status,type,source,description,educationDegree,degreeList,certifications",
    # Availability
    "dateAvailable,willRelocate,desiredLocations,travelLimit",
    # Admin
    "dateAdded,dateLastModified,isDeleted,isAnonymized,onboardingStatus",
    # Custom fields — full range for inspection
    _custom_text,
    _custom_int,
    _custom_float,
    _custom_date,
    _custom_block,
]))

# Rows to skip from the CSV
SKIP_NAMES = {"retainer commencement", "retainer shortlist"}

# Excel styling
HEADER_FILL_COLOR = "FFFFFF00"   # Yellow, matching the Jaybro template
HEADERS    = ["Name", "Current Company", "Role Title", "Location", "Salary", "LI Profile", "Notes"]
COL_WIDTHS = [25,      28,               35,           18,         14,       45,            30]

TIMEOUT = 30
session = requests.Session()
session.headers["User-Agent"] = "futureyou-talent-map/1.0"
session.headers["Accept"]     = "application/json"

# ── Bullhorn Auth ─────────────────────────────────────────────────────────────
def _discover_swimlane(username):
    r = session.get(
        "https://rest.bullhornstaffing.com/rest-services/loginInfo",
        params={"username": username},
        timeout=TIMEOUT,
    )
    r.raise_for_status()
    info = r.json()
    oauth_base = f"{urlsplit(info['oauthUrl']).scheme}://{urlsplit(info['oauthUrl']).netloc}"
    rest_base  = f"{urlsplit(info['restUrl']).scheme}://{urlsplit(info['restUrl']).netloc}"
    return oauth_base, rest_base


def _refresh_access_token(oauth_base, refresh_token):
    resp = session.post(
        f"{oauth_base}/oauth/token",
        data={
            "grant_type":    "refresh_token",
            "refresh_token": refresh_token,
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "redirect_uri":  REDIRECT_URI,
        },
        timeout=TIMEOUT,
    )
    resp.raise_for_status()
    t = resp.json()
    t["_obtained_at"] = int(time.time())
    return t


def _rest_login(rest_base, access_token):
    r = session.get(
        f"{rest_base}/rest-services/login",
        params={"version": "2.0", "access_token": access_token},
        timeout=TIMEOUT,
    )
    r.raise_for_status()
    j = r.json()
    if "BhRestToken" not in j or "restUrl" not in j:
        raise RuntimeError(f"REST login failed: {j}")
    return j


def get_session_creds():
    refresh_token = get_github_variable("BULLHORN_REFRESH_TOKEN_FUTUREYOU")
    if not refresh_token:
        raise SystemExit("Missing GitHub variable: BULLHORN_REFRESH_TOKEN_FUTUREYOU")

    oauth_base, rest_base = _discover_swimlane(USERNAME)
    tokens        = _refresh_access_token(oauth_base, refresh_token)
    access_token  = tokens["access_token"]
    new_refresh   = tokens.get("refresh_token") or refresh_token

    rl = _rest_login(rest_base, access_token)

    if new_refresh != refresh_token:
        update_github_variable("BULLHORN_REFRESH_TOKEN_FUTUREYOU", new_refresh)

    return {"BhRestToken": rl["BhRestToken"], "restUrl": rl["restUrl"]}


# ── Bullhorn Candidate Fetch ───────────────────────────────────────────────────
def get_candidate(rest_url, bh_token, candidate_id):
    r = session.get(
        f"{rest_url}entity/Candidate/{candidate_id}",
        params={"fields": CANDIDATE_FIELDS, "BhRestToken": bh_token},
        timeout=TIMEOUT,
    )
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.json().get("data")


def extract_location(data):
    addr  = data.get("address") or {}
    city  = (addr.get("city")  or "").strip()
    state = (addr.get("state") or "").strip()
    return ", ".join(p for p in [city, state] if p)


def extract_salary(data):
    raw = data.get("salary")
    if not raw:
        return ""
    try:
        return f"${int(raw):,}"
    except (ValueError, TypeError):
        return str(raw)


def find_linkedin(data):
    """Return a full LinkedIn profile URL from the candidate record.

    Priority:
      1. linkedInProfileName — standard Bullhorn field populated by RSC integration
      2. LINKEDIN_FIELD — a customText field already identified in a prior run
      3. Auto-scan customText1–10 for any value containing 'linkedin.com'
    """
    global LINKEDIN_FIELD

    # 1. Standard RSC field — most reliable when LinkedIn Recruiter is connected
    raw = (data.get("linkedInProfileName") or "").strip()
    if raw:
        # Normalise: if it's just a handle/slug (no URL), prepend the base URL
        if raw.startswith("http"):
            return raw
        return f"https://www.linkedin.com/in/{raw.lstrip('/')}"

    # 2. Previously identified custom field
    if LINKEDIN_FIELD:
        return (data.get(LINKEDIN_FIELD) or "").strip()

    # 3. Auto-scan customText1–10
    for i in range(1, 41):
        val = (data.get(f"customText{i}") or "").strip()
        if "linkedin.com" in val.lower():
            LINKEDIN_FIELD = f"customText{i}"
            print(f"  [LinkedIn field auto-discovered: {LINKEDIN_FIELD}]")
            return val

    return ""


# ── CSV Parsing ────────────────────────────────────────────────────────────────
def read_csv(path):
    """Return list of dicts from the Bullhorn Longlist CSV.

    Skips the second header row (field reference names) and Retainer placeholders.
    """
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # The CSV has a second row that mirrors the header with lowercase field paths
            if row.get("ID", "").strip() == "id":
                continue
            name = row.get("Candidate", "").strip()
            if not name or name.lower() in SKIP_NAMES:
                continue
            rows.append(row)
    return rows


# ── Status Sort ────────────────────────────────────────────────────────────────
_STATUS_ORDER = {
    "placed":                   0,
    "1st interview scheduled":  1,
    "submitted":                2,
}

def _sort_key(row):
    return _STATUS_ORDER.get(row.get("Status", "").lower(), 99)


# ── Excel Builder ──────────────────────────────────────────────────────────────
def build_excel(candidates):
    """Build and return an openpyxl Workbook from a list of candidate dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, c in enumerate(candidates, start=2):
        ws.cell(row=row_idx, column=1, value=c["name"])
        ws.cell(row=row_idx, column=2, value=c["company"])
        ws.cell(row=row_idx, column=3, value=c["role"])
        ws.cell(row=row_idx, column=4, value=c["location"])
        ws.cell(row=row_idx, column=5, value=c["salary"])

        li_url = c.get("linkedin", "").strip()
        if li_url:
            cell           = ws.cell(row=row_idx, column=6, value=li_url)
            cell.hyperlink = li_url
            cell.style     = "Hyperlink"
        else:
            ws.cell(row=row_idx, column=6, value="")

        ws.cell(row=row_idx, column=7, value="")  # Notes — left for recruiter to fill in

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Generate Talent Map Excel files from a Bullhorn Longlist CSV export."
    )
    parser.add_argument(
        "--csv",
        default=os.path.join(os.path.dirname(__file__), "rawData.csv"),
        help="Path to Bullhorn Longlist CSV (default: rawData.csv alongside this script)",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(os.path.dirname(__file__), "output"),
        help="Output directory for Excel files (default: ./output/)",
    )
    parser.add_argument(
        "--no-api",
        action="store_true",
        help="Skip Bullhorn API enrichment — useful for testing without credentials",
    )
    args = parser.parse_args()

    print(f"Reading: {args.csv}")
    rows = read_csv(args.csv)
    print(f"  {len(rows)} candidates found (after filtering placeholders)")

    creds = None
    if not args.no_api:
        print("\nAuthenticating with Bullhorn...")
        creds = get_session_creds()
        print(f"  OK — restUrl: {creds['restUrl']}")

    os.makedirs(args.output, exist_ok=True)
    date_str = datetime.today().strftime("%b%y")   # e.g., "Mar26"

    # Use client corp name from the first real row for the filename
    client_corp = (
        rows[0].get("JobOrder.clientCorporation", "Client")
        .replace("/", "-")
        .replace(" ", "")
    ) if rows else "Client"

    print()
    api_dump = []  # collected when --dump-api is set
    candidates = []
    for row in sorted(rows, key=_sort_key):
        cid     = row.get("ID", "").strip()
        name    = row.get("Candidate", "").strip()
        company = row.get("Candidate.companyName", "").strip()
        role    = row.get("Candidate.occupation", "").strip()
        location = salary = linkedin = ""

        if creds and cid:
            try:
                data = get_candidate(creds["restUrl"], creds["BhRestToken"], cid)
                if data:
                    if not args.no_api:
                        api_dump.append({"_candidate_name": name, "_candidate_id": cid, **data})
                    location = extract_location(data)
                    salary   = extract_salary(data)
                    linkedin = find_linkedin(data)
                    print(f"  ✓ {name:<35} loc={location or '—':<20} salary={salary or '—':<10} li={'✓' if linkedin else '—'}")
                else:
                    print(f"  ? {name} — ID {cid} not found via API")
            except Exception as e:
                print(f"  ! {name} — API error: {e}")

        candidates.append({
            "name":     name,
            "company":  company,
            "role":     role,
            "location": location,
            "salary":   salary,
            "linkedin": linkedin,
        })

    wb       = build_excel(candidates)
    filename = f"FYTalentMap_{client_corp}_{date_str}.xlsx"
    out_path = os.path.join(args.output, filename)
    wb.save(out_path)
    print(f"  → Saved: {out_path}")

    if api_dump:
        import json
        dump_path = os.path.join(args.output, "api_dump.json")
        with open(dump_path, "w", encoding="utf-8") as f:
            json.dump(api_dump, f, indent=2, default=str)
        print(f"  → API dump: {dump_path}  ({len(api_dump)} candidates)")

    # LinkedIn field discovery report
    if LINKEDIN_FIELD:
        print(f"\n[LinkedIn field identified: '{LINKEDIN_FIELD}']")
        print(f"  To skip auto-discovery in future runs, set:")
        print(f"  LINKEDIN_FIELD = \"{LINKEDIN_FIELD}\"  (line ~26 of this script)")
    elif not args.no_api and rows:
        print("\n[No LinkedIn URL found in customText1–10 for any candidate]")
        print("  Ask your Bullhorn admin which field stores LinkedIn profile URLs.")
        print("  Then set LINKEDIN_FIELD = 'customTextN' at the top of this script.")

    print("\nDone.")


if __name__ == "__main__":
    main()
