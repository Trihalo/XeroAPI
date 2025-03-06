import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment
import logging
from atbAnalysis import getAtbData

logging.basicConfig(level=logging.INFO)

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Constants
DATE_FORMAT = "%d/%m/%Y"
EXCEL_FILE_DIR = "Overdue"
EXCEL_FILE_PREFIX = "Overdue"
EXCEL_FILE_EXTENSION = ".xlsx"

def processOverdueData(data, client_tokens):
    invoices, overdue_invoices = getAtbData(data, client_tokens)

    df = pd.DataFrame(overdue_invoices)
    df["Inv. Date"] = pd.to_datetime(df["Inv. Date"], format=DATE_FORMAT)
    df = df.sort_values(by="Inv. Date", ascending=True)
    df["Inv. Date"] = df["Inv. Date"].dt.strftime(DATE_FORMAT)

    # file name
    date_string = datetime.today().strftime("%d-%m-%Y")
    os.makedirs(EXCEL_FILE_DIR, exist_ok=True)
    output_file = os.path.join(EXCEL_FILE_DIR, f"{EXCEL_FILE_PREFIX} {date_string}{EXCEL_FILE_EXTENSION}")
    
    df.to_excel(output_file, index=False, startrow=0)

    wb = load_workbook(output_file)
    ws = wb.active

    last_row = len(df) + 1

    # formatting 
    accounting_format = NamedStyle(name="accounting_format")
    accounting_format.number_format = '"$"#,##0.00'

    date_format = NamedStyle(name="date_format")
    date_format.number_format = "DD/MM/YYYY"

    wrap_alignment = Alignment(wrap_text=True)
    wb.add_named_style(accounting_format)
    wb.add_named_style(date_format)
    
    for row in range(1, last_row + 1):
        if row != 2: 
            ws[f"B{row}"].style = "date_format"
            ws[f"C{row}"].style = "date_format"
            ws[f"E{row}"].style = "accounting_format"
            ws[f"G{row}"].alignment = wrap_alignment 
            ws[f"H{row}"].alignment = wrap_alignment 

    column_widths = {"A": 60, "B": 12, "C": 12, "D": 15, "E": 15, "F": 12, "G": 30, "H": 70}

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_file)
    logging.info(f"📂 Excel file '{output_file}' created successfully.")

    return output_file
